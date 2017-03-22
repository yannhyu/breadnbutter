# test_with_sample_cases.py

from openpyxl import load_workbook

FIELD_NAMES = ('payer_id', 'npi', 'match_key', 'ser_date', 'service_date',
               'first', 'middle', 'last', 'dob', 'gender', 
               'social', 'city', 'state', 'zip', 'score',
               'cust_id', 'site_tpg', 'site_ver', 'subs_id', 'batch_id')

def get_acct_data_sets():
    results = []
    wb = load_workbook(filename='test_cases_CH_realtime.xlsx', read_only=True)
    first_sheet = wb.get_sheet_names()[0]
    ws = wb[first_sheet]
    #print('num of cols: {}'.format(ws.max_column))
    #print('num of rows: {}'.format(ws.max_row))

    for row in ws.iter_rows(row_offset=1):    # skip first acct for it is header in excel sheet
        row_scrubbed = [ str(cell.value).strip() for cell in row ]
        row_out = dict(zip(FIELD_NAMES, row_scrubbed))
        results.append(row_out)
    return results


if __name__ == '__main__':
    accts = get_acct_data_sets()
    for acct in accts:
        #print(acct)
        print('|'.join([acct.get(key) for key in FIELD_NAMES]))
